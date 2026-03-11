"""コンプライアンスレポート自動生成 テストスイート"""

import io
from datetime import UTC, datetime
from unittest.mock import AsyncMock, patch

from src.services.report_generator import (
    ComplianceReportBuilder,
    ExcelReportGenerator,
    PdfReportGenerator,
    ReportData,
)

# ── ReportData テスト ──────────────────────────────────────────────────────────


class TestReportData:
    def test_create_report_data(self):
        data = ReportData(
            title="テストレポート",
            period="2026年3月",
            generated_at=datetime.now(UTC),
            sections=[],
        )
        assert data.title == "テストレポート"
        assert data.period == "2026年3月"
        assert data.sections == []


# ── ComplianceReportBuilder テスト ────────────────────────────────────────────


class TestComplianceReportBuilder:
    def test_build_jsox_change_report_empty(self):
        data = ComplianceReportBuilder.build_jsox_change_report([], "2026年3月")
        assert "J-SOX" in data.title
        assert data.period == "2026年3月"
        assert len(data.sections) == 1
        assert "0件" in data.sections[0]["summary"]

    def test_build_jsox_change_report_with_data(self):
        changes = [
            {
                "change_number": "CHG-2026-000001",
                "title": "テスト変更",
                "change_type": "Normal",
                "status": "Completed",
                "risk_level": "Low",
                "created_at": "2026-03-01",
            }
        ]
        data = ComplianceReportBuilder.build_jsox_change_report(changes, "2026年3月")
        assert len(data.sections[0]["rows"]) == 1
        assert data.sections[0]["rows"][0][0] == "CHG-2026-000001"

    def test_build_incident_analysis_report(self):
        incidents = [
            {
                "incident_number": "INC-2026-000001",
                "title": "サーバーダウン",
                "priority": "P1",
                "status": "Resolved",
                "sla_breached": True,
                "created_at": "2026-03-01",
            },
            {
                "incident_number": "INC-2026-000002",
                "title": "軽微障害",
                "priority": "P4",
                "status": "Closed",
                "sla_breached": False,
                "created_at": "2026-03-02",
            },
        ]
        data = ComplianceReportBuilder.build_incident_analysis_report(incidents, "2026年3月")
        assert "インシデント" in data.title
        assert "P1: 1件" in data.sections[0]["summary"]
        assert "SLA違反: 1件" in data.sections[0]["summary"]

    def test_build_cmdb_inventory_report(self):
        items = [
            {
                "ci_name": "WebServer01", "ci_type": "Server",
                "status": "Active", "environment": "Production", "owner": "admin",
            }
        ]
        data = ComplianceReportBuilder.build_cmdb_inventory_report(items, "2026年3月")
        assert "CMDB" in data.title
        assert len(data.sections[0]["rows"]) == 1

    def test_build_audit_trail_report(self):
        logs = [
            {
                "action": "LOGIN",
                "entity_type": "User",
                "user_id": "user-123",
                "ip_address": "192.168.1.1",
                "created_at": "2026-03-01 10:00",
            }
        ]
        data = ComplianceReportBuilder.build_audit_trail_report(logs, "2026年3月")
        assert "監査" in data.title
        assert len(data.sections[0]["rows"]) == 1

    def test_generated_at_defaults_to_now(self):
        before = datetime.now(UTC)
        data = ComplianceReportBuilder.build_jsox_change_report([], "2026年3月")
        after = datetime.now(UTC)
        assert before <= data.generated_at <= after

    def test_custom_generated_at(self):
        ts = datetime(2026, 1, 1, tzinfo=UTC)
        data = ComplianceReportBuilder.build_jsox_change_report([], "2026年1月", generated_at=ts)
        assert data.generated_at == ts


# ── Excel 生成テスト ─────────────────────────────────────────────────────────


class TestExcelReportGenerator:
    def test_generate_returns_bytes(self):
        gen = ExcelReportGenerator()
        data = ReportData(
            title="テスト",
            period="2026年3月",
            generated_at=datetime.now(UTC),
            sections=[
                {
                    "name": "シート1",
                    "summary": "テスト用サマリー",
                    "headers": ["列A", "列B", "列C"],
                    "rows": [["val1", "val2", "val3"]],
                }
            ],
        )
        content = gen.generate(data)
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_generate_valid_xlsx(self):
        """生成バイト列が有効なXLSXかどうか確認"""
        import openpyxl

        gen = ExcelReportGenerator()
        data = ComplianceReportBuilder.build_jsox_change_report(
            [
                {
                    "change_number": "CHG-001",
                    "title": "test",
                    "change_type": "Normal",
                    "status": "Completed",
                    "risk_level": "Low",
                    "created_at": "2026-03-01",
                }
            ],
            "2026年3月",
        )
        content = gen.generate(data)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        assert "サマリー" in wb.sheetnames
        assert "変更管理一覧" in wb.sheetnames

    def test_generate_empty_sections(self):
        gen = ExcelReportGenerator()
        data = ReportData(
            title="空テスト",
            period="2026年3月",
            generated_at=datetime.now(UTC),
            sections=[],
        )
        content = gen.generate(data)
        assert len(content) > 0

    def test_generate_long_sheet_name_truncated(self):
        """シート名が31文字超でも正常動作するか確認"""
        import openpyxl

        gen = ExcelReportGenerator()
        data = ReportData(
            title="テスト",
            period="2026年3月",
            generated_at=datetime.now(UTC),
            sections=[
                {
                    "name": "A" * 50,  # 50文字のシート名
                    "headers": ["列1"],
                    "rows": [["値1"]],
                }
            ],
        )
        content = gen.generate(data)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        # 31文字に切り詰めされているはず
        assert any(len(name) <= 31 for name in wb.sheetnames)


# ── PDF 生成テスト ────────────────────────────────────────────────────────────


class TestPdfReportGenerator:
    def test_generate_returns_bytes(self):
        gen = PdfReportGenerator()
        data = ReportData(
            title="テストPDFレポート",
            period="2026年3月",
            generated_at=datetime.now(UTC),
            sections=[
                {
                    "name": "テストセクション",
                    "summary": "テスト用サマリー",
                    "headers": ["Col1", "Col2"],
                    "rows": [["val1", "val2"]],
                }
            ],
        )
        content = gen.generate(data)
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_generate_valid_pdf(self):
        """生成バイト列が PDF シグネチャで始まるか確認"""
        gen = PdfReportGenerator()
        data = ComplianceReportBuilder.build_incident_analysis_report([], "2026年3月")
        content = gen.generate(data)
        assert content[:4] == b"%PDF"

    def test_generate_empty_sections(self):
        gen = PdfReportGenerator()
        data = ReportData(
            title="空PDF",
            period="2026年3月",
            generated_at=datetime.now(UTC),
            sections=[],
        )
        content = gen.generate(data)
        assert content[:4] == b"%PDF"


# ── API エンドポイント テスト ─────────────────────────────────────────────────


class TestComplianceReportsAPI:
    def setup_method(self):
        from fastapi.testclient import TestClient

        from src.main import app

        self.client = TestClient(app, raise_server_exceptions=False)

    def test_list_report_types(self):
        """レポート種別一覧API"""
        resp = self.client.get("/api/v1/compliance-reports/types")
        assert resp.status_code == 200
        data = resp.json()
        assert len(data) == 4
        types = {r["type"] for r in data}
        assert "jsox_change" in types
        assert "incident_analysis" in types

    def test_preview_report_no_db(self):
        """プレビューAPI（DBなし環境でも動作）"""
        with patch(
            "src.api.v1.compliance_reports._build_jsox_change",
            new=AsyncMock(
                return_value=ComplianceReportBuilder.build_jsox_change_report([], "2026年3月")
            ),
        ):
            resp = self.client.get(
                "/api/v1/compliance-reports/preview",
                params={"report_type": "jsox_change", "year": 2026, "month": 3},
            )
        assert resp.status_code == 200
        data = resp.json()
        assert data["title"] == "J-SOX 変更管理レポート"
        assert "sections" in data

    def test_generate_excel_report(self):
        """Excel レポート生成API"""
        with patch(
            "src.api.v1.compliance_reports._build_incident_analysis",
            new=AsyncMock(
                return_value=ComplianceReportBuilder.build_incident_analysis_report([], "2026年3月")
            ),
        ):
            resp = self.client.get(
                "/api/v1/compliance-reports/generate",
                params={
                    "report_type": "incident_analysis",
                    "output_format": "excel",
                    "year": 2026,
                    "month": 3,
                },
            )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert len(resp.content) > 0

    def test_generate_pdf_report(self):
        """PDF レポート生成API"""
        with patch(
            "src.api.v1.compliance_reports._build_jsox_change",
            new=AsyncMock(
                return_value=ComplianceReportBuilder.build_jsox_change_report([], "2026年3月")
            ),
        ):
            resp = self.client.get(
                "/api/v1/compliance-reports/generate",
                params={
                    "report_type": "jsox_change",
                    "output_format": "pdf",
                    "year": 2026,
                    "month": 3,
                },
            )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/pdf"
        assert resp.content[:4] == b"%PDF"

    def test_generate_content_disposition(self):
        """ダウンロードヘッダーの確認"""
        with patch(
            "src.api.v1.compliance_reports._build_cmdb_inventory",
            new=AsyncMock(
                return_value=ComplianceReportBuilder.build_cmdb_inventory_report([], "2026年3月")
            ),
        ):
            resp = self.client.get(
                "/api/v1/compliance-reports/generate",
                params={
                    "report_type": "cmdb_inventory",
                    "output_format": "excel",
                    "year": 2026,
                    "month": 3,
                },
            )
        assert resp.status_code == 200
        assert "content-disposition" in resp.headers
        assert "cmdb_inventory_202603.xlsx" in resp.headers["content-disposition"]
