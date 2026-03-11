"""コンプライアンスレポート生成エンジン - PDF/Excel"""

import io
from datetime import UTC, datetime
from typing import Any


class ReportData:
    """レポートデータ格納クラス"""

    def __init__(
        self,
        title: str,
        period: str,
        generated_at: datetime,
        sections: list[dict[str, Any]],
    ) -> None:
        self.title = title
        self.period = period
        self.generated_at = generated_at
        self.sections = sections


class ExcelReportGenerator:
    """openpyxl ベースの Excel レポート生成"""

    def generate(self, data: ReportData) -> bytes:
        """レポートデータから Excel バイト列を生成"""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        # サマリーシート
        ws = wb.active
        ws.title = "サマリー"

        # タイトル行
        ws["A1"] = data.title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"対象期間: {data.period}"
        ws["A3"] = f"生成日時: {data.generated_at.strftime('%Y-%m-%d %H:%M:%S')} UTC"
        ws["A4"] = ""

        # セクション毎にシートを生成
        for section in data.sections:
            ws_sec = wb.create_sheet(title=section["name"][:31])  # Excel シート名最大31文字
            ws_sec["A1"] = section["name"]
            ws_sec["A1"].font = Font(bold=True, size=12)

            rows = section.get("rows", [])
            headers = section.get("headers", [])

            if headers:
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws_sec.cell(row=3, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(rows, start=4):
                for col_idx, value in enumerate(row, start=1):
                    ws_sec.cell(row=row_idx, column=col_idx, value=value)

            # 列幅自動調整
            for col in ws_sec.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws_sec.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            # サマリーにセクション概要を追記
            summary_row = ws.max_row + 2
            ws.cell(row=summary_row, column=1, value=section["name"]).font = Font(bold=True)
            if "summary" in section:
                ws.cell(row=summary_row + 1, column=1, value=section["summary"])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


class PdfReportGenerator:
    """reportlab ベースの PDF レポート生成"""

    def generate(self, data: ReportData) -> bytes:
        """レポートデータから PDF バイト列を生成"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )
        styles = getSampleStyleSheet()
        story = []

        # タイトル
        story.append(Paragraph(data.title, styles["Title"]))
        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph(f"対象期間: {data.period}", styles["Normal"]))
        story.append(
            Paragraph(
                f"生成日時: {data.generated_at.strftime('%Y-%m-%d %H:%M:%S')} UTC",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 10 * mm))

        # セクション
        for section in data.sections:
            story.append(Paragraph(section["name"], styles["Heading2"]))
            if "summary" in section:
                story.append(Paragraph(section["summary"], styles["Normal"]))
                story.append(Spacer(1, 4 * mm))

            rows = section.get("rows", [])
            headers = section.get("headers", [])
            if headers and rows:
                table_data = [headers] + [list(r) for r in rows]
                tbl = Table(table_data, repeatRows=1)
                tbl.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            (
                                "ROWBACKGROUNDS",
                                (0, 1),
                                (-1, -1),
                                [colors.white, colors.HexColor("#EEF2FF")],
                            ),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("TOPPADDING", (0, 0), (-1, -1), 4),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ]
                    )
                )
                story.append(tbl)
                story.append(Spacer(1, 6 * mm))

        doc.build(story)
        return buf.getvalue()


# ── レポートテンプレート ────────────────────────────────────────────────────────


class ComplianceReportBuilder:
    """コンプライアンスレポートデータ構築"""

    @staticmethod
    def build_jsox_change_report(
        changes: list[dict],
        period: str,
        generated_at: datetime | None = None,
    ) -> ReportData:
        """J-SOX 変更管理レポートデータを構築"""
        if generated_at is None:
            generated_at = datetime.now(UTC)
        rows = [
            [
                c.get("change_number", ""),
                c.get("title", ""),
                c.get("change_type", ""),
                c.get("status", ""),
                c.get("risk_level", ""),
                c.get("created_at", ""),
            ]
            for c in changes
        ]
        return ReportData(
            title="J-SOX 変更管理レポート",
            period=period,
            generated_at=generated_at,
            sections=[
                {
                    "name": "変更管理一覧",
                    "summary": f"対象期間の変更件数: {len(changes)}件",
                    "headers": [
                        "変更番号", "タイトル", "変更種別", "ステータス", "リスクレベル", "作成日時"
                    ],
                    "rows": rows,
                }
            ],
        )

    @staticmethod
    def build_incident_analysis_report(
        incidents: list[dict],
        period: str,
        generated_at: datetime | None = None,
    ) -> ReportData:
        """インシデント分析レポートデータを構築"""
        if generated_at is None:
            generated_at = datetime.now(UTC)
        p1 = [i for i in incidents if i.get("priority") == "P1"]
        breached = [i for i in incidents if i.get("sla_breached")]
        rows = [
            [
                i.get("incident_number", ""),
                i.get("title", ""),
                i.get("priority", ""),
                i.get("status", ""),
                "あり" if i.get("sla_breached") else "なし",
                i.get("created_at", ""),
            ]
            for i in incidents
        ]
        summary = (
            f"総件数: {len(incidents)}件 | P1: {len(p1)}件 | SLA違反: {len(breached)}件"
        )
        return ReportData(
            title="インシデント分析レポート",
            period=period,
            generated_at=generated_at,
            sections=[
                {
                    "name": "インシデント一覧",
                    "summary": summary,
                    "headers": [
                        "インシデント番号", "タイトル", "優先度",
                        "ステータス", "SLA違反", "作成日時",
                    ],
                    "rows": rows,
                }
            ],
        )

    @staticmethod
    def build_cmdb_inventory_report(
        items: list[dict],
        period: str,
        generated_at: datetime | None = None,
    ) -> ReportData:
        """CMDB 資産台帳レポートデータを構築"""
        if generated_at is None:
            generated_at = datetime.now(UTC)
        rows = [
            [
                i.get("ci_name", ""),
                i.get("ci_type", ""),
                i.get("status", ""),
                i.get("environment", ""),
                i.get("owner", ""),
            ]
            for i in items
        ]
        return ReportData(
            title="CMDB 資産台帳",
            period=period,
            generated_at=generated_at,
            sections=[
                {
                    "name": "構成アイテム一覧",
                    "summary": f"総資産数: {len(items)}件",
                    "headers": ["CI名", "種別", "ステータス", "環境", "オーナー"],
                    "rows": rows,
                }
            ],
        )

    @staticmethod
    def build_audit_trail_report(
        logs: list[dict],
        period: str,
        generated_at: datetime | None = None,
    ) -> ReportData:
        """セキュリティ監査証跡レポートデータを構築"""
        if generated_at is None:
            generated_at = datetime.now(UTC)
        rows = [
            [
                log.get("action", ""),
                log.get("entity_type", ""),
                log.get("user_id", ""),
                log.get("ip_address", ""),
                log.get("created_at", ""),
            ]
            for log in logs
        ]
        return ReportData(
            title="セキュリティ監査証跡レポート",
            period=period,
            generated_at=generated_at,
            sections=[
                {
                    "name": "監査ログ",
                    "summary": f"ログ件数: {len(logs)}件",
                    "headers": [
                        "アクション", "エンティティ種別", "ユーザーID", "IPアドレス", "日時"
                    ],
                    "rows": rows,
                }
            ],
        )
